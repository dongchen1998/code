#!usr/bin/python
import numpy as np
import pandas as pd
import os
import sys
import openpyxl
import matplotlib
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.comments import Comment
import matplotlib.gridspec as gridspec
from matplotlib.pyplot import MultipleLocator
from openpyxl.styles import PatternFill,colors,Border,Side


def eachFile(file,sheet,shuru,fudong,writer):
   array1 = [["ND" for j in range(0, 48)] for i in range(0, 32)]
   array2 = [["ND" for j in range(0, 24)] for i in range(0, 16)]
   array3 = [["ND" for j in range(0, 12)] for i in range(0, 8)]
   #读取文件夹文件列表
   pathDir = os.listdir(file)
   l=len(pathDir)
   array=[]
   if(l<=96):
      array=array3
   if((l<=384)&(l>96)):
      array=array2
   if((l<=1536)&(l>384)):
      array=array1
   row_class = {'A':0,'B':1,'C':2,'D':3,'E':4,'F':5,'G':6,'H':7,'I':8,'J':9,'K':10,'L':11,'M':12,'N':13,'O':14,'P':15,'Q':16,'R':17,'S':18,'T':19,'U':20,'V':21,'W':22,'X':23,'Y':24,'Z':25,'AA':26,'AB':27,'AC':28,'AD':29,'AE':30,'AF':31}
   for allDir in pathDir:
      row1 = allDir[4]
      row2 = allDir[5]
      col1 = allDir[5]
      col2 = allDir[6]
      col3 = allDir[7]
      col4 = allDir[8]
        #判断文件名格式并拆分
      if (row2 =="A")|(row2 =="B")|(row2 =="C")|(row2 =="D")|(row2 =="E")|(row2 =="F"):
         row = str(row1)+str(row2)
         row = int(row_class.get(row,[]))
         if col3 != "_":
            col = int(str(col2)+str(col3))-1
         else:
            col = int(col2)-1
      else:
         row = int(row_class.get(row1,[]))
         if col2 != "_":
            col = int(str(col1)+str(col2))-1
         else:
            col = int(col1)-1
    #print(row,col)  #row是字母，col是数字
      fopen = open(file+'\\'+allDir,'r')
      for line in fopen.readlines():
         line = line.strip()
         if line[0]!="C":
            line = line.split('\t')
            value = line[1] #value是第二列的值
            line1 = float(line[0])
            key=float("%.1f"%line1)
            if (float(shuru)<=(key+float(fudong)))&(float(shuru)>=(key-float(fudong))):
               array[row][col]=value
    
   data = pd.DataFrame(array)
   data.to_excel(writer, sheet, float_format='%.5f')
	# ‘page_1’是写入excel的sheet名
    #writer.save()
    #return array


def solution():
   #lujing= input("文件夹路径: ")
   shuru1 = input("Targeted m/z: ") #要提取的
   shuru2 = input("IS m/z: ") #要提取的
   fudong = input("Tolerance: ") #变化范围
   label = input("lable: ")
   file1 = input("Import: ") #run
   #file2 = input("同位素内标响应值: ") #run2
   result_name = input("Export: ")
   result=result_name+'.xlsx'
   fig=result_name+'.png'
   writer = pd.ExcelWriter(result)		# 写入Excel文件
   document_list = os.listdir(file1)

    
   pathDir=eachFile(file1,"Targeted",shuru1,fudong,writer)
   pathDir=eachFile(file1,"IS",shuru2,fudong,writer)
   writer.save()
   #开始对导入数据进行处理
   wb = openpyxl.load_workbook(result)
   wb.create_sheet("Targeted_IS")
   wb.create_sheet("Lable",0)
   wb.create_sheet("Average")
   wb.create_sheet("SD")
   w1 = wb["Targeted"]
   w2 = wb["IS"]
   w3 = wb["Targeted_IS"]
   w4 = wb["Lable"]
   w5 = wb["Average"]
   w6 = wb["SD"]
   #行数，列数，x,y轴标签
   aaa=0;bbb=0;ll=0;xlab=[];ylab=[]
   l=len(document_list)
   #判断数据大小
   if(l<=96):
     aaa=10;bbb=14;ll=96
     picture=plt.figure(figsize =([8, 5]))
   if((l<=384)&(l>96)):
     aaa=18;bbb=26;ll=384
     picture=plt.figure(figsize =([8, 5]))
   if((l<=1536)&(l>384)):
     aaa=34;bbb=50;ll=1536
     picture=plt.figure(figsize =([16, 10]))
   xlab=[i for i in range(1,bbb-1)]
   for i in range(2,aaa):
     if i<=27:
        ylab.append(chr(i+63))
     else:
        ylab.append(chr(i-26+63))
   ylab=ylab[::-1]
   ylab1=[i for i in range(1,aaa-1)]
   
   #浓度计算
   vmax=0
   for i in range(2,aaa):
      for j in range(2,bbb):
         if (w1.cell(row=i,column=j).value != "ND") & (w2.cell(row=i,column=j).value != "ND"):
            w3.cell(row=i,column=j).value=float(float(w1.cell(row=i,column=j).value))/(float(w2.cell(row=i,column=j).value))
            if(w3.cell(row=i,column=j).value>vmax):
                vmax=w3.cell(row=i,column=j).value
         if (w1.cell(row=i,column=j).value == "ND") & (w2.cell(row=i,column=j).value != "ND"):
            w3.cell(row=i,column=j).value="ND"
         if (w1.cell(row=i,column=j).value != "ND") & (w2.cell(row=i,column=j).value == "ND"):
            w3.cell(row=i,column=j).value="ISND"
            w2.cell(row=i,column=j).value="ISND"
         if (w1.cell(row=i,column=j).value == "ND") & (w2.cell(row=i,column=j).value == "ND"):
            w2.cell(row=i,column=j).value="ISND"
            w3.cell(row=i,column=j).value="ISND"
   #如果有平行
   pingxing1,pingxing2 = input("m n: ").split()
   a=2;b=2
   vmax_ave=vmax_sd=0
   for i in range(0,(aaa-2)//(int(pingxing1))):
      for j in range(0,(bbb-2)//(int(pingxing2))):
         Sum=0
         c=[]
         num_ND=num_ISND= 0
         for hang in range(a+i*int(pingxing1),int(a+(i+1)*int(pingxing1))):
            for lie in range(b+j*int(pingxing2),int(b+(j+1)*int(pingxing2))):
               
               if (w3.cell(hang,lie).value != "ND")&(w3.cell(hang,lie).value != "ISND")&(w3.cell(hang,lie).value != None):
                  Sum += float(w3.cell(hang,lie).value)
                  c.append(float(w3.cell(hang,lie).value))
               if(w3.cell(hang,lie).value == "ND"):
                  num_ND=num_ND+1
               if(w3.cell(hang,lie).value == "ISND"):
                  num_ISND=num_ISND+1
         if(Sum!=0):
            std = np.std(c)
            ave = np.mean(c)
            d1 = str(round(ave,6))
            d2 = str(round(std/ave,6))
            w5.cell(row=i+2,column=j+2).value=d1
            w6.cell(row=i+2,column=j+2).value=d2
            if(str(w6.cell(row=i+2,column=j+2).value)=="0.0"):
              w6.cell(row=i+2,column=j+2).value=0
            if(vmax_ave<float(d1)):
                vmax_ave=float(d1)
            if(vmax_sd<float(d2)):
                vmax_sd=float(d2)
         if(Sum==0):
            if(num_ND==int(pingxing1)*int(pingxing2)):
                w5.cell(row=i+2,column=j+2).value="ND"
                w6.cell(row=i+2,column=j+2).value="ND"
            if(num_ISND==int(pingxing1)*int(pingxing2)):
                w5.cell(row=i+2,column=j+2).value="ISND"
                w6.cell(row=i+2,column=j+2).value="ISND"
            else:
                w5.cell(row=i+2,column=j+2).value="ND"
                w6.cell(row=i+2,column=j+2).value="ND" 
   f = open(label,'r')
   x = 1                #在excel开始写的位置（y）
   y = 1
   while True:  #循环，读取文本里面的所有内容
      line = f.readline() #一行一行读取
      if not line:  #如果没有内容，则退出循环
         break
      for i in line.split('\t'):#读取出相应的内容写到x
         item=i.strip()
         w4.cell(row=x,column=y).value=item
         y += 1 #另起一列
      x += 1 #另起一行
      y = 1  #初始成第一列
   f.close()
   #准备填充的颜色
   c=np.linspace(0,vmax,11)
   c_ave=np.linspace(0,vmax_ave,11)
   c_sd=np.linspace(0,vmax_sd,11)
   fill0 = PatternFill("solid", fgColor="1F4E79")
   fill1 = PatternFill("solid", fgColor="2E75B6")
   fill2 = PatternFill("solid", fgColor="7AA6E0")
   fill3 = PatternFill("solid", fgColor="BDD7EE")
   fill4 = PatternFill("solid", fgColor="C9D1DD")
   fill5 = PatternFill("solid", fgColor="E1CBC9")
   fill6 = PatternFill("solid", fgColor="D89E9F")
   fill7 = PatternFill("solid", fgColor="DB676A")
   fill8 = PatternFill("solid", fgColor="D8181C")
   fill9 = PatternFill("solid", fgColor="A40000")
   column_dict={1:'A',2:'B',3:'C',4:'D',5:'E',6:'F',7:'G',8:'H',9:'I',10:'J',11:'K',12:'L',13:'M',14:'N',15:'O',16:'P',17:'Q',18:'R',19:'S',20:'T',21:'U',22:'V',23:'W',24:'X',25:'Y',26:'Z',27:'AA',28:'AB',29:'AC',30:'AD',31:'AE',32:'AF',33:'AG',34:'AH',35:'AI',36:'AJ',37:'AK',38:'AL',39:'AM',40:'AN',41:'AO',42:'AP',43:'AQ',44:'AR',45:'AS',46:'AT',47:'AU',48:'AV',49:'AW'}
   #set border
   border_set = Border(left=Side(style="thin", color="C8C8C8"),
                    right=Side(style="thin", color="C8C8C8"),
                    top=Side(style="thin", color="C8C8C8"),
                    bottom=Side(style="thin", color="C8C8C8"))
   #给浓度进行填色
   for i in range(2,aaa):
      for j in  range(2,bbb):
         pos = column_dict.get(j,[])+str(i)
         comments = Comment(w4.cell(row=i,column=j).value,'zzd')
         comments.width = 150
         comments.height = 40
         w3[pos].comment=comments
         if (w3[pos].value!="ND")&(w3[pos].value!="ISND"):
            if ((w3[pos].value)>=c[0]) & ((w3[pos].value)<c[1]):
               w3[pos].fill = fill0
            if ((w3[pos].value)>c[1]) & ((w3[pos].value)<c[2]):
               w3[pos].fill = fill1
            if ((w3[pos].value)>c[2]) & ((w3[pos].value)<c[3]):
               w3[pos].fill = fill2
            if ((w3[pos].value)>c[3]) & ((w3[pos].value)<c[4]):
               w3[pos].fill = fill3
            if ((w3[pos].value)>c[4]) & ((w3[pos].value)<c[5]):
               w3[pos].fill = fill4
            if ((w3[pos].value)>c[5]) & ((w3[pos].value)<c[6]):
               w3[pos].fill = fill5
            if ((w3[pos].value)>c[6]) & ((w3[pos].value)<c[7]):
               w3[pos].fill = fill6
            if ((w3[pos].value)>c[7]) & ((w3[pos].value)<c[8]):
               w3[pos].fill = fill7
            if ((w3[pos].value)>c[8]) & ((w3[pos].value)<c[9]):
               w3[pos].fill = fill8
            if ((w3[pos].value)>c[9]) & ((w3[pos].value)<=c[10]):
               w3[pos].fill = fill9
         if(w3[pos].value=="ND"):
            w3[pos].fill = fill0
         if(w3[pos].value=="ISND"):
             w3[pos].fill=PatternFill("solid", fgColor="ffffff")
   for i in range(2,((aaa-2)//int(pingxing1)+2)):
      for j in  range(2,((bbb-2)//int(pingxing2)+2)):
         pos = column_dict.get(j,[])+str(i)
         if (w5[pos].value!="ND")&(w5[pos].value!="ISND"):
            if (float(w5[pos].value)>=c_ave[0]) & (float(w5[pos].value)<c_ave[1]):
               w5[pos].fill = fill0
            if (float(w5[pos].value)>c_ave[1]) & (float(w5[pos].value)<c_ave[2]):
               w5[pos].fill = fill1
            if (float(w5[pos].value)>c_ave[2]) & (float(w5[pos].value)<c_ave[3]):
               w5[pos].fill = fill2
            if (float(w5[pos].value)>c_ave[3]) & (float(w5[pos].value)<c_ave[4]):
               w5[pos].fill = fill3
            if (float(w5[pos].value)>c_ave[4]) & (float(w5[pos].value)<c_ave[5]):
               w5[pos].fill = fill4
            if (float(w5[pos].value)>c_ave[5]) & (float(w5[pos].value)<c_ave[6]):
               w5[pos].fill = fill5
            if (float(w5[pos].value)>c_ave[6]) & (float(w5[pos].value)<c_ave[7]):
               w5[pos].fill = fill6
            if (float(w5[pos].value)>c_ave[7]) & (float(w5[pos].value)<c_ave[8]):
               w5[pos].fill = fill7
            if (float(w5[pos].value)>c_ave[8]) & (float(w5[pos].value)<c_ave[9]):
               w5[pos].fill = fill8
            if (float(w5[pos].value)>c_ave[9]) & (float(w5[pos].value)<=c_ave[10]):
               w5[pos].fill = fill9
         if(w5[pos].value=="ND"):
            w5[pos].fill = fill0
         if(w5[pos].value=="ISND"):
             w5[pos].fill=PatternFill("solid", fgColor="ffffff")
   #给单元格加边框
   for i in range(1,aaa):
      for j in  range(1,bbb):
         pos = column_dict.get(j,[])+str(i)
         w1[pos].border = border_set
         w2[pos].border = border_set
         w3[pos].border = border_set
         w4[pos].border = border_set
         w5[pos].border = border_set
   for i in range(2,aaa):
      if i<=27:
         w1.cell(row=i,column=1).value=chr(i+63)
         w2.cell(row=i,column=1).value=chr(i+63)
         w3.cell(row=i,column=1).value=chr(i+63)
      else:
         w1.cell(row=i,column=1).value=chr(65)+chr(i-26+63)
         w2.cell(row=i,column=1).value=chr(65)+chr(i-26+63)
         w3.cell(row=i,column=1).value=chr(65)+chr(i-26+63)
   for i in range(2,((aaa-2)//int(pingxing1)+2)):
      if i<=27:
         w5.cell(row=i,column=1).value=chr(i+63)
         w6.cell(row=i,column=1).value=chr(i+63)
      else:
         w5.cell(row=i,column=1).value=chr(65)+chr(i-26+63)
         w6.cell(row=i,column=1).value=chr(65)+chr(i-26+63)
   for j in range(2,bbb):
      w1.cell(row=1,column=j).value=j-1
      w2.cell(row=1,column=j).value=j-1
      w3.cell(row=1,column=j).value=j-1
   for j in range(2,((bbb-2)//int(pingxing2)+2)):
      w5.cell(row=1,column=j).value=j-1
      w6.cell(row=1,column=j).value=j-1
   
   #画图
   X=[0 for x in range(0,ll)]
   Y=[0 for x in range(0,ll)]
   pos=0
   for i in range(1,(aaa-1)):
       for j in range(1,(bbb-1)):
           X[pos]=i;Y[pos]=j
           pos=pos+1
   
   norm = matplotlib.colors.Normalize(vmin=0.00, vmax=vmax)
   #define colors
   colors = ["#1F4E79","#2E75B6","#7AA6E0","#BDD7EE","#D3D9E3","#E1CBC9","#D89E9F","#DB676A","#D8181C","#A40000"]
   cmap= matplotlib.colors.ListedColormap(colors)
   cmap.set_under("crimson")
   cmap.set_over("w")
   #gs = gridspec.GridSpec(1,28)
   #ax1 = plt.subplot(gs[0,:26])
   #ax2 = plt.subplot(gs[0,27:28])
   for i in range(0,ll):
       x=X[i];y=Y[i]
       pos = column_dict.get((y+1),[])+str(x+1)
       if (w3[pos].value=="ND"):
           a=plt.scatter((y),(aaa-x-1),c=0.00,s=100,cmap = cmap,norm = norm,alpha=1)
       
       if (w3[pos].value!="ND")&(w3[pos].value!="ISND"):
           a=plt.scatter((y),(aaa-x-1),c=w3[pos].value,s=100,cmap = cmap,norm = norm,alpha=1)
   plt.colorbar(a)
   for i in range(0,ll):
       x=X[i];y=Y[i]
       pos = column_dict.get((y+1),[])+str(x+1)
       if (w3[pos].value=="ISND"):
           a=plt.scatter((y),(aaa-x-1),c="#ffffff",s=100)
       
   #plt.show()
   #plt.close()
   #y_major_locator=MultipleLocator(1)
#把y轴的刻度间隔设置为10，并存在变量里
   ax=plt.gca()
#ax为两条坐标轴的实例
   #ax.yaxis.set_major_locator(y_major_locator)
   
   ax.set_xticks(xlab)
   ax.xaxis.set_ticks_position('top')
   ax.set_yticks(ylab1)
   ax.set_yticklabels(ylab)

   

   plt.show()  
   #plt.yticks(ylab1,ylab)
   picture.savefig(result_name+'1'+'.png')

   ##picture2
   #画图
   picture1=plt.figure(figsize =([8, 5]))
   xlab=ylab=[]
   xlab=[i for i in range(1,(bbb-2)//int(pingxing2)+1)]
   for i in range(2,(aaa-2)//int(pingxing1)+2):
     if i<=27:
        ylab.append(chr(i+63))
     else:
        ylab.append(chr(i-26+63))
   ylab=ylab[::-1]
   ylab1=[i for i in range(1,(aaa-2)//int(pingxing1)+1)]
   X1=[0 for x in range(0,ll//(int(pingxing1)*int(pingxing2)))]
   Y1=[0 for x in range(0,ll//(int(pingxing1)*int(pingxing2)))]
   pos=0
   for i in range(1,((aaa-2)//int(pingxing1))+1):
       for j in range(1,((bbb-2)//int(pingxing2))+1):
           X1[pos]=i;Y1[pos]=j
           pos=pos+1
   #print(vmax_ave)
   norm1 = matplotlib.colors.Normalize(vmin=0.00, vmax=vmax_ave)
   #define colors
   colors = ["#1F4E79","#2E75B6","#7AA6E0","#BDD7EE","#D3D9E3","#E1CBC9","#D89E9F","#DB676A","#D8181C","#A40000"]
   cmap= matplotlib.colors.ListedColormap(colors)
   cmap.set_under("crimson")
   cmap.set_over("w")
   for i in range(0,ll//(int(pingxing1)*int(pingxing2))):
       x=X1[i];y=Y1[i]
       pos = column_dict.get((y+1),[])+str(x+1)
       if (str(w6[pos].value)=="0"):
           a=plt.scatter((y),((aaa-2)//int(pingxing1)-x+1),c=float(w5[pos].value),s=0.01*500,cmap = cmap,norm = norm1,alpha=1)
       if (w5[pos].value!="ND")&(w5[pos].value!="ISND"):
           a=plt.scatter((y),((aaa-2)//int(pingxing1)-x+1),c=float(w5[pos].value),s=float(w6[pos].value or 0)*500,cmap = cmap,norm = norm1,alpha=1)
   plt.colorbar(a)
       
   ax1=plt.gca()
   
   ax1.set_xticks(xlab)
   ax1.xaxis.set_ticks_position('top')
   ax1.set_yticks(ylab1)
   ax1.set_yticklabels(ylab)
   plt.show()  
   #plt.yticks(ylab1,ylab)
   picture1.savefig(result_name+'2'+'.png')


   wb.save(result)
   print("Finished!")

def main():
   #print('WELCOME! I am YuChang')
   flag = True
   #temp = 0
   while flag:
     #  temp = temp + 1
     #  print(temp)
     #  test()
      solution()
      print('输入q,即可结束程序;输入Enter,程序重新执行')
      num = input("请输入： ")
      if num == 'q':
         sys.exit(0)

if __name__ == '__main__':
   main()

